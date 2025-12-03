"""
DuckDB-based Excel processor for efficient Excel file parsing.

DuckDB provides fast, in-process analytical queries that work excellently
for processing Excel data, especially for large files.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from od_parse.utils.logging_utils import get_logger

# Optional third-party dependencies
DUCKDB_AVAILABLE: bool = False
OPENPYXL_AVAILABLE: bool = False

try:
    import duckdb

    DUCKDB_AVAILABLE = True
except ImportError:
    duckdb = None  # type: ignore[assignment]

try:
    import openpyxl  # noqa: F401

    OPENPYXL_AVAILABLE = True
except ImportError:
    pass

logger = get_logger(__name__)


class OutputFormat(Enum):
    """Output format options for Excel processing."""

    JSON = "json"
    MARKDOWN = "markdown"


@dataclass
class ExcelProcessorConfig:
    """Configuration for Excel processing."""

    # Output format: 'json' or 'markdown'
    output_format: OutputFormat = OutputFormat.JSON

    # Sheet selection: None means all sheets
    sheets: Optional[List[str]] = None

    # Whether to include sheet names as keys in output
    include_sheet_names: bool = True

    # Whether to treat first row as headers
    first_row_as_header: bool = True

    # Skip empty rows
    skip_empty_rows: bool = True

    # Skip empty columns (columns where all values are null/empty)
    skip_empty_columns: bool = True

    # Skip unnamed columns (pandas default "Unnamed: X" columns)
    skip_unnamed_columns: bool = True

    # Skip columns where all values are null
    skip_all_null_columns: bool = True

    # Maximum rows to process per sheet (None = no limit)
    max_rows: Optional[int] = None

    # SQL query to run on the data (optional)
    sql_query: Optional[str] = None

    # Column name mapping (rename columns)
    column_mapping: Dict[str, str] = field(default_factory=dict)

    # Data type casting
    type_casting: Dict[str, str] = field(default_factory=dict)


class ExcelProcessor:
    """
    DuckDB-based Excel processor for efficient data extraction.

    Supports:
    - Excel files (.xlsx, .xls)
    - Multiple sheets
    - SQL queries on data
    - JSON and Markdown output
    - Large file handling
    """

    def __init__(self, config: Optional[ExcelProcessorConfig] = None):
        """
        Initialize the Excel processor.

        Args:
            config: Optional configuration for processing
        """
        if not DUCKDB_AVAILABLE:
            raise ImportError(
                "DuckDB is required for Excel processing. "
                "Install with: pip install od-parse[excel]"
            )

        if not OPENPYXL_AVAILABLE:
            logger.warning(
                "openpyxl not installed. Some Excel features may be limited. "
                "Install with: pip install openpyxl"
            )

        self.config = config or ExcelProcessorConfig()
        self.conn = duckdb.connect(":memory:")

        # Install and load spatial extension for better data handling
        try:
            self.conn.execute("INSTALL spatial;")
            self.conn.execute("LOAD spatial;")
        except Exception:
            pass  # Extension not critical

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()

    def close(self):
        """Close the DuckDB connection."""
        if self.conn:
            self.conn.close()

    def parse(
        self,
        file_path: Union[str, Path],
        output_format: Optional[OutputFormat] = None,
        **kwargs,
    ) -> Union[Dict[str, Any], str]:
        """
        Parse an Excel file and return data in the specified format.

        Args:
            file_path: Path to the Excel file
            output_format: Override config output format
            **kwargs: Additional options to override config

        Returns:
            Parsed data as dict (JSON) or string (Markdown)
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if file_path.suffix.lower() not in [".xlsx", ".xls", ".xlsm", ".xlsb"]:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")

        # Determine output format
        fmt = output_format or self.config.output_format
        if isinstance(fmt, str):
            fmt = OutputFormat(fmt.lower())

        # Load Excel data
        sheets_data = self._load_excel(file_path, **kwargs)

        # Apply transformations
        sheets_data = self._apply_transformations(sheets_data, **kwargs)

        # Convert to output format
        if fmt == OutputFormat.JSON:
            return self._to_json(sheets_data, file_path.name)
        else:
            return self._to_markdown(sheets_data, file_path.name)

    def _load_excel(self, file_path: Path, **kwargs) -> Dict[str, List[Dict[str, Any]]]:
        """
        Load Excel file into DuckDB and extract data.

        Args:
            file_path: Path to Excel file

        Returns:
            Dictionary of sheet names to row data
        """
        sheets_data = {}

        try:
            # Get sheet names
            sheets = self._get_sheet_names(file_path)
            target_sheets = kwargs.get("sheets") or self.config.sheets or sheets

            for sheet_name in target_sheets:
                if sheet_name not in sheets:
                    logger.warning(f"Sheet '{sheet_name}' not found in {file_path}")
                    continue

                try:
                    # Read sheet using DuckDB or pandas fallback
                    result = self._load_sheet(file_path, sheet_name)

                    if result is not None:
                        # Apply row/column filtering
                        if self.config.skip_empty_rows:
                            result = result.dropna(how="all")

                        if self.config.skip_empty_columns:
                            result = result.dropna(axis=1, how="all")

                        # Skip unnamed columns (pandas default "Unnamed: X")
                        if self.config.skip_unnamed_columns:
                            unnamed_cols = [
                                col
                                for col in result.columns
                                if str(col).startswith("Unnamed:")
                            ]
                            if unnamed_cols:
                                result = result.drop(columns=unnamed_cols)
                                logger.debug(
                                    f"Dropped {len(unnamed_cols)} unnamed columns"
                                )

                        # Skip columns where all values are null
                        if self.config.skip_all_null_columns:
                            null_cols = [
                                col
                                for col in result.columns
                                if result[col].isna().all()
                            ]
                            if null_cols:
                                result = result.drop(columns=null_cols)
                                logger.debug(
                                    f"Dropped {len(null_cols)} all-null columns"
                                )

                        if self.config.max_rows:
                            result = result.head(self.config.max_rows)

                        sheets_data[sheet_name] = result.to_dict("records")

                except Exception as e:
                    logger.error(f"Error reading sheet '{sheet_name}': {e}")
                    sheets_data[sheet_name] = []

        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise

        return sheets_data

    def _load_sheet(self, file_path: Path, sheet_name: str):
        """Load a single sheet from Excel file."""
        try:
            # Try DuckDB's read_xlsx_auto first (if spatial extension loaded)
            query = f"""
                SELECT * FROM read_xlsx_auto(
                    '{str(file_path)}',
                    sheet = '{sheet_name}',
                    header = {str(self.config.first_row_as_header).lower()}
                )
            """
            return self.conn.execute(query).fetchdf()
        except Exception:
            # Fallback to pandas
            return self._load_with_pandas(file_path, sheet_name)

    def _get_sheet_names(self, file_path: Path) -> List[str]:
        """Get list of sheet names from Excel file."""
        try:
            # Try openpyxl first
            if OPENPYXL_AVAILABLE:
                from openpyxl import load_workbook

                wb = load_workbook(file_path, read_only=True, data_only=True)
                return wb.sheetnames
        except Exception:
            pass

        # Fallback to pandas
        try:
            import pandas as pd

            xl = pd.ExcelFile(file_path)
            return xl.sheet_names
        except Exception as e:
            logger.error(f"Could not get sheet names: {e}")
            return ["Sheet1"]

    def _load_with_pandas(self, file_path: Path, sheet_name: str):
        """Fallback method to load Excel using pandas."""
        try:
            import pandas as pd

            return pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=0 if self.config.first_row_as_header else None,
            )
        except ImportError:
            raise ImportError(
                "pandas is required for Excel processing. "
                "Install with: pip install pandas openpyxl"
            )

    def _apply_transformations(
        self, sheets_data: Dict[str, List[Dict[str, Any]]], **kwargs
    ) -> Dict[str, List[Dict[str, Any]]]:
        """Apply configured transformations to the data."""
        column_mapping = kwargs.get("column_mapping") or self.config.column_mapping
        sql_query = kwargs.get("sql_query") or self.config.sql_query

        transformed = {}

        for sheet_name, rows in sheets_data.items():
            if not rows:
                transformed[sheet_name] = []
                continue

            # Apply column mapping
            if column_mapping:
                rows = [
                    {column_mapping.get(k, k): v for k, v in row.items()}
                    for row in rows
                ]

            # Apply SQL query if provided
            if sql_query:
                rows = self._apply_sql_query(rows, sheet_name, sql_query)

            transformed[sheet_name] = rows

        return transformed

    def _apply_sql_query(
        self, rows: List[Dict[str, Any]], sheet_name: str, sql_query: str
    ) -> List[Dict[str, Any]]:
        """Apply SQL query to sheet data using DuckDB."""
        if not rows:
            return rows

        try:
            import pandas as pd

            # Create temporary table
            df = pd.DataFrame(rows)
            table_name = "sheet_data"

            self.conn.register(table_name, df)

            # Execute query (replace 'sheet' placeholder with actual table)
            query = sql_query.replace("{sheet}", table_name)
            result = self.conn.execute(query).fetchdf()

            return result.to_dict("records")

        except Exception as e:
            logger.error(f"SQL query error on sheet '{sheet_name}': {e}")
            return rows

    def _to_json(
        self, sheets_data: Dict[str, List[Dict[str, Any]]], filename: str
    ) -> Dict[str, Any]:
        """Convert sheet data to JSON-compatible dictionary."""
        result = {
            "filename": filename,
            "sheet_count": len(sheets_data),
            "sheets": {},
        }

        for sheet_name, rows in sheets_data.items():
            # Clean data for JSON serialization
            clean_rows = []
            for row in rows:
                clean_row = {}
                for k, v in row.items():
                    # Handle NaN, NaT, etc.
                    if v is None or (hasattr(v, "__class__") and v != v):  # NaN check
                        clean_row[str(k)] = None
                    elif hasattr(v, "isoformat"):  # datetime
                        clean_row[str(k)] = v.isoformat()
                    else:
                        clean_row[str(k)] = v
                clean_rows.append(clean_row)

            result["sheets"][sheet_name] = {
                "row_count": len(clean_rows),
                "columns": list(clean_rows[0].keys()) if clean_rows else [],
                "data": clean_rows,
            }

        return result

    def _to_markdown(
        self, sheets_data: Dict[str, List[Dict[str, Any]]], filename: str
    ) -> str:
        """Convert sheet data to Markdown format."""
        markdown = []
        markdown.append(f"# {filename}\n")

        for sheet_name, rows in sheets_data.items():
            markdown.append(f"## {sheet_name}\n")

            if not rows:
                markdown.append("*No data in this sheet*\n")
                continue

            # Get headers
            headers = list(rows[0].keys())

            # Create table header
            markdown.append("| " + " | ".join(str(h) for h in headers) + " |")
            markdown.append("| " + " | ".join(["---"] * len(headers)) + " |")

            # Add rows
            for row in rows:
                values = []
                for h in headers:
                    val = row.get(h, "")
                    # Handle None and NaN
                    if val is None or (hasattr(val, "__class__") and val != val):
                        val = ""
                    # Escape pipe characters in cell values
                    val = str(val).replace("|", "\\|")
                    values.append(val)
                markdown.append("| " + " | ".join(values) + " |")

            markdown.append("")  # Empty line after table

        return "\n".join(markdown)

    def query(self, file_path: Union[str, Path], sql: str) -> List[Dict[str, Any]]:
        """
        Run a SQL query on Excel data.

        Args:
            file_path: Path to Excel file
            sql: SQL query to execute

        Returns:
            Query results as list of dictionaries
        """
        # Load all sheets
        file_path = Path(file_path)
        sheets_data = self._load_excel(file_path)

        try:
            import pandas as pd

            # Register all sheets as tables
            for sheet_name, rows in sheets_data.items():
                if rows:
                    table_name = sheet_name.replace(" ", "_").replace("-", "_")
                    df = pd.DataFrame(rows)
                    self.conn.register(table_name, df)

            # Execute query
            result = self.conn.execute(sql).fetchdf()
            return result.to_dict("records")

        except Exception as e:
            logger.error(f"Query error: {e}")
            raise

    def get_sheet_info(self, file_path: Union[str, Path]) -> Dict[str, Dict[str, Any]]:
        """
        Get information about sheets in an Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            Dictionary with sheet information
        """
        file_path = Path(file_path)
        sheets_data = self._load_excel(file_path)

        info = {}
        for sheet_name, rows in sheets_data.items():
            info[sheet_name] = {
                "row_count": len(rows),
                "columns": list(rows[0].keys()) if rows else [],
                "column_count": len(rows[0]) if rows else 0,
            }

        return info


# Convenience functions


def parse_excel(
    file_path: Union[str, Path],
    output_format: str = "json",
    sheets: Optional[List[str]] = None,
    first_row_as_header: bool = True,
    sql_query: Optional[str] = None,
    **kwargs,
) -> Union[Dict[str, Any], str]:
    """
    Parse an Excel file and return data in the specified format.

    Args:
        file_path: Path to the Excel file
        output_format: 'json' or 'markdown'
        sheets: List of sheet names to process (None = all)
        first_row_as_header: Treat first row as headers
        sql_query: Optional SQL query to run on data
        **kwargs: Additional configuration options

    Returns:
        Parsed data as dict (JSON) or string (Markdown)

    Example:
        >>> # Parse to JSON
        >>> data = parse_excel("data.xlsx", output_format="json")
        >>> print(json.dumps(data, indent=2))

        >>> # Parse to Markdown
        >>> md = parse_excel("data.xlsx", output_format="markdown")
        >>> print(md)

        >>> # With SQL query
        >>> data = parse_excel(
        ...     "data.xlsx",
        ...     sql_query="SELECT * FROM {sheet} WHERE amount > 100"
        ... )
    """
    config = ExcelProcessorConfig(
        output_format=OutputFormat(output_format.lower()),
        sheets=sheets,
        first_row_as_header=first_row_as_header,
        sql_query=sql_query,
    )

    with ExcelProcessor(config) as processor:
        return processor.parse(file_path)


def excel_to_json(
    file_path: Union[str, Path],
    sheets: Optional[List[str]] = None,
    output_file: Optional[Union[str, Path]] = None,
    **kwargs,
) -> Dict[str, Any]:
    """
    Convert Excel file to JSON format.

    Args:
        file_path: Path to the Excel file
        sheets: List of sheet names to process (None = all)
        output_file: Optional path to save JSON output
        **kwargs: Additional configuration options

    Returns:
        Parsed data as dictionary

    Example:
        >>> data = excel_to_json("report.xlsx")
        >>> for sheet_name, sheet_data in data["sheets"].items():
        ...     print(f"{sheet_name}: {sheet_data['row_count']} rows")
    """
    result = parse_excel(file_path, output_format="json", sheets=sheets, **kwargs)

    if output_file:
        output_path = Path(output_file)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, default=str)
        logger.info(f"JSON saved to {output_path}")

    return result


def excel_to_markdown(
    file_path: Union[str, Path],
    sheets: Optional[List[str]] = None,
    output_file: Optional[Union[str, Path]] = None,
    **kwargs,
) -> str:
    """
    Convert Excel file to Markdown format.

    Args:
        file_path: Path to the Excel file
        sheets: List of sheet names to process (None = all)
        output_file: Optional path to save Markdown output
        **kwargs: Additional configuration options

    Returns:
        Markdown formatted string

    Example:
        >>> md = excel_to_markdown("report.xlsx")
        >>> print(md)

        >>> # Save to file
        >>> excel_to_markdown("report.xlsx", output_file="report.md")
    """
    result = parse_excel(file_path, output_format="markdown", sheets=sheets, **kwargs)

    if output_file:
        output_path = Path(output_file)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(result)
        logger.info(f"Markdown saved to {output_path}")

    return result


def check_duckdb_available() -> bool:
    """Check if DuckDB is available for Excel processing."""
    return DUCKDB_AVAILABLE
